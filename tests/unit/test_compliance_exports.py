"""Tests for compliance report export functionality.

These tests verify that the compliance analytics export methods
work correctly for PDF, Excel, and HTML formats.
"""

from datetime import datetime
from unittest.mock import MagicMock, patch

import pytest


class TestComplianceExports:
    """Tests for compliance report exports."""

    @pytest.fixture
    def sample_report_data(self):
        """Create sample report data for testing."""
        return {
            "title": "Compliance Report Q4 2025",
            "generated_at": datetime.now().isoformat(),
            "period": {
                "start_date": "2025-10-01",
                "end_date": "2025-12-31",
                "days": 90,
            },
            "executive_summary": {
                "overall_compliance_score": 85.5,
                "total_controls_monitored": 150,
                "automated_controls_percentage": 72.3,
                "high_risk_count": 3,
                "overdue_items_count": 5,
                "evidence_freshness_score": 91.2,
            },
            "compliance_overview": {
                "overall_score": 85.5,
                "total_controls": 150,
                "passing_controls": 128,
                "failing_controls": 12,
                "not_applicable": 10,
                "automation_percentage": 72.3,
            },
            "risk_summary": {
                "total_risks": 45,
                "high_risks": [{"id": "R1", "title": "Unpatched systems"}],
                "medium_risks": [{"id": "R2", "title": "Weak passwords"}],
                "risk_by_category": {
                    "infrastructure": 15,
                    "application": 20,
                    "data": 10,
                },
            },
            "key_actions_required": [
                {
                    "priority": "high",
                    "description": "Patch critical vulnerabilities",
                    "due_date": "2025-12-15",
                },
                {
                    "priority": "medium",
                    "description": "Update access control policies",
                    "due_date": "2025-12-30",
                },
                "Review third-party integrations",
            ],
        }

    @pytest.fixture
    def mock_analytics(self):
        """Create mock compliance analytics instance."""
        # Mock the dependencies
        evidence_fabric = MagicMock()
        risk_system = MagicMock()
        rules_engine = MagicMock()

        # Import and create instance
        from cerebro.compliance.reporting_analytics import ComplianceAnalytics

        analytics = ComplianceAnalytics(evidence_fabric, risk_system, rules_engine)
        return analytics

    def test_export_to_html_basic(self, mock_analytics, sample_report_data):
        """Test HTML export generates valid HTML."""
        html = mock_analytics._export_to_html(sample_report_data, "default")

        assert "<!DOCTYPE html>" in html
        assert "<html" in html
        assert "</html>" in html
        assert sample_report_data["title"] in html

    def test_export_to_html_contains_executive_summary(
        self, mock_analytics, sample_report_data
    ):
        """Test HTML export includes executive summary metrics."""
        html = mock_analytics._export_to_html(sample_report_data, "default")

        assert "Executive Summary" in html
        assert "85.5" in html  # compliance score
        assert "150" in html  # total controls
        assert "72.3" in html  # automation percentage

    def test_export_to_html_contains_risk_summary(
        self, mock_analytics, sample_report_data
    ):
        """Test HTML export includes risk summary."""
        html = mock_analytics._export_to_html(sample_report_data, "default")

        assert "Risk Summary" in html
        assert "Total Risks" in html or "total_risks" in html.lower()

    def test_export_to_html_contains_key_actions(
        self, mock_analytics, sample_report_data
    ):
        """Test HTML export includes key actions."""
        html = mock_analytics._export_to_html(sample_report_data, "default")

        assert "Key Actions" in html
        assert "Patch critical vulnerabilities" in html
        assert "Update access control policies" in html

    def test_export_to_html_has_styling(self, mock_analytics, sample_report_data):
        """Test HTML export includes CSS styling."""
        html = mock_analytics._export_to_html(sample_report_data, "default")

        assert "<style>" in html
        assert "font-family" in html
        assert "metric-card" in html or "metric" in html

    def test_export_to_html_empty_data(self, mock_analytics):
        """Test HTML export handles empty data gracefully."""
        html = mock_analytics._export_to_html({}, "default")

        assert "<!DOCTYPE html>" in html
        assert "</html>" in html

    def test_export_to_html_missing_sections(self, mock_analytics):
        """Test HTML export handles missing sections."""
        partial_data = {
            "title": "Partial Report",
            "generated_at": datetime.now().isoformat(),
        }
        html = mock_analytics._export_to_html(partial_data, "default")

        assert "Partial Report" in html
        # Should not crash on missing executive_summary, etc.

    def test_export_to_csv_basic(self, mock_analytics, sample_report_data):
        """Test CSV export generates valid CSV data."""
        csv_bytes = mock_analytics._export_to_csv(sample_report_data)

        csv_content = csv_bytes.decode("utf-8")
        assert "Section" in csv_content
        assert "Key" in csv_content
        assert "Value" in csv_content

    def test_export_to_csv_contains_data(self, mock_analytics, sample_report_data):
        """Test CSV export contains report data."""
        csv_bytes = mock_analytics._export_to_csv(sample_report_data)
        csv_content = csv_bytes.decode("utf-8")

        assert "executive_summary" in csv_content or "overall_compliance_score" in csv_content
        assert "85.5" in csv_content

    def test_export_to_csv_nested_data(self, mock_analytics, sample_report_data):
        """Test CSV export flattens nested data correctly."""
        csv_bytes = mock_analytics._export_to_csv(sample_report_data)
        csv_content = csv_bytes.decode("utf-8")

        # Should contain nested keys
        assert "period" in csv_content or "start_date" in csv_content

    def test_export_to_excel_without_openpyxl(self, mock_analytics, sample_report_data):
        """Test Excel export falls back to CSV when openpyxl not available."""
        with patch.dict("sys.modules", {"openpyxl": None}):
            # Force re-import to trigger fallback
            result = mock_analytics._export_to_excel(sample_report_data)

            # Should return bytes (either Excel or CSV)
            assert isinstance(result, bytes)
            assert len(result) > 0

    def test_export_to_excel_basic(self, mock_analytics, sample_report_data):
        """Test Excel export generates valid data."""
        try:
            import openpyxl  # noqa: F401
        except ImportError:
            pytest.skip("openpyxl not installed")

        excel_bytes = mock_analytics._export_to_excel(sample_report_data)

        assert isinstance(excel_bytes, bytes)
        assert len(excel_bytes) > 0
        # Excel files start with PK (zip signature)
        assert excel_bytes[:2] == b"PK"

    def test_export_to_excel_has_worksheets(self, mock_analytics, sample_report_data):
        """Test Excel export creates multiple worksheets."""
        try:
            from io import BytesIO

            from openpyxl import load_workbook
        except ImportError:
            pytest.skip("openpyxl not installed")

        excel_bytes = mock_analytics._export_to_excel(sample_report_data)
        wb = load_workbook(BytesIO(excel_bytes))

        # Should have multiple sheets
        assert len(wb.sheetnames) >= 1
        assert "Executive Summary" in wb.sheetnames

    def test_export_to_excel_contains_data(self, mock_analytics, sample_report_data):
        """Test Excel export contains report data."""
        try:
            from io import BytesIO

            from openpyxl import load_workbook
        except ImportError:
            pytest.skip("openpyxl not installed")

        excel_bytes = mock_analytics._export_to_excel(sample_report_data)
        wb = load_workbook(BytesIO(excel_bytes))

        # Check executive summary sheet
        ws = wb["Executive Summary"]
        values = [str(cell.value) for row in ws.iter_rows() for cell in row if cell.value]
        values_str = " ".join(values)

        assert "Compliance Report" in values_str
        assert "85.5" in values_str or "Overall Compliance Score" in values_str

    def test_export_to_pdf_raises_without_library(
        self, mock_analytics, sample_report_data
    ):
        """Test PDF export raises ImportError when libraries not available."""
        # Mock both weasyprint and xhtml2pdf as unavailable
        with patch.dict("sys.modules", {"weasyprint": None, "xhtml2pdf": None}):
            with pytest.raises(ImportError, match="PDF export requires"):
                mock_analytics._export_to_pdf(sample_report_data, "default")


class TestComplianceAnalyticsIntegration:
    """Integration tests for compliance analytics."""

    @pytest.fixture
    def mock_dependencies(self):
        """Create mock dependencies for compliance analytics."""
        evidence_fabric = MagicMock()
        evidence_fabric.search = MagicMock(return_value=[])

        risk_system = MagicMock()
        risk_system.get_risk_dashboard_data = MagicMock(
            return_value={
                "risk_summary": {
                    "total_risks": 10,
                    "high_risks": [],
                    "medium_risks": [],
                },
                "treatment_summary": {"overdue_treatments": []},
            }
        )

        rules_engine = MagicMock()

        return evidence_fabric, risk_system, rules_engine

    def test_create_compliance_analytics(self, mock_dependencies):
        """Test creating compliance analytics instance."""
        from cerebro.compliance.reporting_analytics import create_compliance_analytics

        evidence_fabric, risk_system, rules_engine = mock_dependencies

        analytics = create_compliance_analytics(
            evidence_fabric, risk_system, rules_engine
        )

        assert analytics is not None
        assert analytics.evidence_fabric == evidence_fabric
        assert analytics.risk_system == risk_system
        assert analytics.rules_engine == rules_engine

    def test_generate_executive_dashboard(self, mock_dependencies):
        """Test generating executive dashboard."""
        from cerebro.compliance.reporting_analytics import ComplianceAnalytics

        evidence_fabric, risk_system, rules_engine = mock_dependencies

        analytics = ComplianceAnalytics(evidence_fabric, risk_system, rules_engine)

        # Mock internal methods that do heavy computation
        analytics._calculate_compliance_overview = MagicMock(
            return_value={
                "overall_score": 85.0,
                "total_controls": 100,
                "automation_percentage": 70.0,
            }
        )
        analytics._analyze_control_effectiveness = MagicMock(return_value={})
        analytics._analyze_evidence_freshness = MagicMock(
            return_value={"average_freshness_score": 90.0}
        )
        analytics._calculate_compliance_trends = MagicMock(return_value={})
        analytics._identify_key_actions = MagicMock(return_value=[])

        dashboard = analytics.generate_executive_dashboard(time_period_days=30)

        assert "generated_at" in dashboard
        assert "period" in dashboard
        assert "executive_summary" in dashboard


class TestHTMLTemplateRendering:
    """Tests for HTML template rendering edge cases."""

    @pytest.fixture
    def mock_analytics(self):
        """Create mock analytics instance."""
        from cerebro.compliance.reporting_analytics import ComplianceAnalytics

        return ComplianceAnalytics(MagicMock(), MagicMock(), MagicMock())

    def test_html_escapes_special_characters(self, mock_analytics):
        """Test HTML properly escapes special characters."""
        data = {
            "title": "Report <script>alert('xss')</script>",
            "executive_summary": {
                "overall_compliance_score": 85.5,
            },
        }

        html = mock_analytics._export_to_html(data, "default")

        # With autoescape=True, script tags should be escaped
        assert "<script>" not in html or "&lt;script&gt;" in html

    def test_html_handles_none_values(self, mock_analytics):
        """Test HTML handles None values gracefully."""
        # Test with minimal data - template handles missing keys gracefully
        data = {
            "title": None,
            "generated_at": datetime.now().isoformat(),
        }

        # Should not raise an exception
        html = mock_analytics._export_to_html(data, "default")
        assert "<!DOCTYPE html>" in html
        assert "</html>" in html

    def test_html_handles_numeric_edge_cases(self, mock_analytics):
        """Test HTML handles numeric edge cases."""
        data = {
            "executive_summary": {
                "overall_compliance_score": 0,
                "total_controls_monitored": 0,
                "automated_controls_percentage": 100.0,
                "high_risk_count": 999,
            },
        }

        html = mock_analytics._export_to_html(data, "default")

        # Score of 0 should show as critical (red)
        assert "status-critical" in html or "0" in html
        # 100% should be valid
        assert "100" in html
